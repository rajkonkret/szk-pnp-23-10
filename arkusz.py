import openpyxl

# pip install openpyxl

# workbook = openpyxl.load_workbook('sales.xlsx')
workbook = openpyxl.load_workbook('C:\\Users\\cscomarch\\PycharmProjects\\szk-pnp-23-10\\sales.xlsx')
worksheet = workbook.active

print(worksheet)  # <Worksheet "Arkusz1">

lista = []

for w in worksheet:  # (<Cell 'Arkusz1'.A1>, <Cell 'Arkusz1'.B1>, <Cell 'Arkusz1'.C1>)
    print(w)

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        lista.append(col[i].value)

print(lista)
print(lista[0:3])

for i in range(0, len(lista), 3):  # start, stop, krok
    print(lista[i:i + 3])
